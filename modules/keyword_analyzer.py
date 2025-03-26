import os
import openpyxl
from collections import Counter
from config import CONFIG
from utils.helpers import create_directory, safe_remove_file
from typing import List, Tuple, Dict, Optional
import logging
from tqdm import tqdm
from datetime import datetime
import re

def setup_logging(log_file: Optional[str] = None) -> None:
    """
    Configure logging for the analysis process
    
    Args:
        log_file (Optional[str]): Path to log file
    """
    log_file = log_file or "logs/keyword_analysis.log"
    
    # Create logs directory if it doesn't exist
    os.makedirs("logs", exist_ok=True)
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler()
        ]
    )

def analyze_keywords(input_file: Optional[str] = None, output_file: Optional[str] = None) -> Tuple[bool, Dict]:
    """
    Enhanced keyword analysis with:
    - Better error handling
    - Config validation
    - Progress tracking
    - Advanced keyword processing
    - Multiple output formats
    
    Args:
        input_file (Optional[str]): Path to input Excel file
        output_file (Optional[str]): Path to output Excel file
    
    Returns:
        Tuple[bool, Dict]: (Success status, Statistics dictionary)
    """
    try:
        # Initialize statistics
        stats = {
            "total_rows": 0,
            "processed_rows": 0,
            "total_keywords": 0,
            "unique_keywords": 0,
            "start_time": datetime.now(),
            "end_time": None
        }
        
        setup_logging()
        logging.info("Starting keyword analysis process")

        # Validate and get configuration with defaults
        config_analysis = CONFIG.get("analysis", {})
        
        input_file = input_file or config_analysis.get("input_file", "output/extracted_data.xlsx")
        output_file = output_file or config_analysis.get("output_file", "output/keyword_analysis.xlsx")
        top_n = config_analysis.get("top_n", 50)
        min_keyword_length = config_analysis.get("min_keyword_length", 3)
        exclude_numbers = config_analysis.get("exclude_numbers", True)
        backup_enabled = config_analysis.get("backup_enabled", True)
        
        # Create output directory if needed
        create_directory(os.path.dirname(output_file))
        logging.info(f"Output will be saved to: {output_file}")

        # Backup existing output file if enabled
        if backup_enabled and os.path.exists(output_file):
            backup_file = f"{output_file}.bak"
            safe_remove_file(backup_file)  # Remove old backup if exists
            os.rename(output_file, backup_file)

        # Load workbook with error handling
        try:
            logging.info(f"Loading input file: {input_file}")
            wb = openpyxl.load_workbook(input_file)
            ws = wb.active
        except FileNotFoundError:
            raise Exception(f"Input file not found: {input_file}")
        except Exception as e:
            raise Exception(f"Error loading workbook: {str(e)}")

        # Validate worksheet
        if ws.max_row < 2:
            logging.warning("No data found to analyze (empty worksheet)")
            return False, stats

        stats["total_rows"] = ws.max_row - 1

        # Extract and process keywords
        logging.info("Extracting keywords from worksheet...")
        all_keywords = extract_all_keywords(ws, min_keyword_length, exclude_numbers, stats)
        
        if not all_keywords:
            logging.warning("No keywords found in the dataset")
            return False, stats

        stats["total_keywords"] = len(all_keywords)
        stats["unique_keywords"] = len(set(all_keywords))

        # Analyze keyword frequencies
        logging.info("Analyzing keyword frequencies...")
        keyword_counts = Counter(all_keywords)
        top_keywords = keyword_counts.most_common(top_n)
        
        # Generate analysis report
        logging.info("Generating analysis report...")
        generate_analysis_report(output_file, top_keywords, stats)
        
        # Additional analysis (optional)
        if config_analysis.get("generate_wordcloud", False):
            generate_wordcloud(keyword_counts, output_file)
        
        stats["end_time"] = datetime.now()
        logging.info(f"Successfully completed analysis. Top {top_n} keywords saved to {output_file}")
        return True, stats

    except Exception as e:
        logging.error(f"Error in keyword analysis: {str(e)}", exc_info=True)
        return False, stats

def extract_all_keywords(worksheet, min_length: int = 3, exclude_numbers: bool = True, stats: Optional[Dict] = None) -> List[str]:
    """
    Extract and preprocess keywords from worksheet
    
    Args:
        worksheet: Excel worksheet to process
        min_length (int): Minimum keyword length
        exclude_numbers (bool): Whether to exclude numeric keywords
        stats (Optional[Dict]): Statistics dictionary to update
    
    Returns:
        List[str]: List of extracted keywords
    """
    keywords = []
    total_rows = worksheet.max_row - 1
    
    # Create progress bar
    pbar = tqdm(range(2, worksheet.max_row + 1), desc="Extracting keywords")
    
    for row in pbar:
        try:
            keywords_cell = worksheet.cell(row=row, column=5).value
            if keywords_cell:
                # Process keywords with additional filtering
                row_keywords = [
                    kw.strip().lower() 
                    for kw in keywords_cell.split(',') 
                    if (kw.strip() and 
                        len(kw.strip()) >= min_length and
                        (not exclude_numbers or not kw.strip().isdigit()) and
                        not re.match(r'^[0-9\s]+$', kw.strip()))  # Exclude pure numbers with spaces
                ]
                keywords.extend(row_keywords)
            
            if stats:
                stats["processed_rows"] += 1
            pbar.set_description(f"Processed {row-1}/{total_rows} rows")
                
        except Exception as e:
            logging.warning(f"Error processing row {row}: {str(e)}")
            continue
    
    return keywords

def generate_analysis_report(output_path: str, keyword_data: List[Tuple[str, int]], stats: Dict) -> None:
    """
    Generate Excel report with keyword analysis
    
    Args:
        output_path (str): Path to save the report
        keyword_data (List[Tuple[str, int]]): List of (keyword, count) tuples
        stats (Dict): Statistics dictionary
    """
    wb = openpyxl.Workbook()
    
    # Main analysis sheet
    ws = wb.active
    ws.title = "Keyword Analysis"
    
    # Header with additional metadata
    ws.append(["Keyword", "Frequency", "Percentage"])
    
    total_occurrences = sum(count for _, count in keyword_data)
    
    for keyword, count in keyword_data:
        percentage = (count / total_occurrences) * 100
        ws.append([keyword, count, f"{percentage:.2f}%"])
    
    # Add summary statistics
    ws.append([])
    ws.append(["Summary Statistics", ""])
    ws.append(["Total rows processed", stats["processed_rows"]])
    ws.append(["Total keywords found", stats["total_keywords"]])
    ws.append(["Unique keywords", stats["unique_keywords"]])
    ws.append(["Total occurrences", total_occurrences])
    ws.append(["Start time", stats["start_time"]])
    ws.append(["End time", stats["end_time"]])
    
    # Formatting
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    
    # Add metadata sheet
    metadata_ws = wb.create_sheet("Metadata")
    metadata_ws.append(["Analysis Parameters"])
    metadata_ws.append(["Input File", os.path.basename(output_path)])
    metadata_ws.append(["Output File", os.path.basename(output_path)])
    metadata_ws.append(["Total Rows", stats["total_rows"]])
    metadata_ws.append(["Processed Rows", stats["processed_rows"]])
    metadata_ws.append(["Total Keywords", stats["total_keywords"]])
    metadata_ws.append(["Unique Keywords", stats["unique_keywords"]])
    metadata_ws.append(["Start Time", stats["start_time"]])
    metadata_ws.append(["End Time", stats["end_time"]])
    
    wb.save(output_path)

def generate_wordcloud(keyword_counts: Dict[str, int], output_path: str) -> None:
    """
    Generate wordcloud visualization
    
    Args:
        keyword_counts (Dict[str, int]): Dictionary of keyword frequencies
        output_path (str): Base path for output files
    """
    try:
        from wordcloud import WordCloud
        import matplotlib.pyplot as plt
        
        logging.info("Generating wordcloud...")
        
        wordcloud = WordCloud(
            width=800,
            height=600,
            background_color='white',
            max_words=100,
            max_font_size=150,
            random_state=42,
            colormap='viridis'
        ).generate_from_frequencies(keyword_counts)
        
        plt.figure(figsize=(10, 8))
        plt.imshow(wordcloud, interpolation='bilinear')
        plt.axis('off')
        
        # Save wordcloud image
        img_path = os.path.splitext(output_path)[0] + '_wordcloud.png'
        plt.savefig(img_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        logging.info(f"Wordcloud saved to: {img_path}")
    except ImportError:
        logging.warning("WordCloud package not available. Skipping wordcloud generation.")
    except Exception as e:
        logging.warning(f"Error generating wordcloud: {str(e)}")

if __name__ == "__main__":
    try:
        success, stats = analyze_keywords()
        if success:
            print("Keyword analysis completed successfully")
            print(f"Processed {stats['processed_rows']}/{stats['total_rows']} rows")
            print(f"Found {stats['total_keywords']} keywords ({stats['unique_keywords']} unique)")
        else:
            print("Keyword analysis failed")
    except Exception as e:
        print(f"Fatal error: {str(e)}")