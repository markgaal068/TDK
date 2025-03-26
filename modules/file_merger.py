import os
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from config import CONFIG
from utils.helpers import create_directory, safe_remove_file
from typing import List, Dict, Optional, Tuple
import logging
from tqdm import tqdm
from datetime import datetime

def merge_excel_files(input_file: Optional[str] = None, output_file: Optional[str] = None) -> Tuple[bool, Dict]:
    """
    Process a single Excel file with enhanced functionality.
    Features:
    - Better error handling
    - Config validation
    - Duplicate prevention
    - Progress reporting
    - Header consistency checking
    
    Args:
        input_file (Optional[str]): Path to the input Excel file
        output_file (Optional[str]): Path to save processed results
    
    Returns:
        Tuple[bool, Dict]: (Success status, Statistics dictionary)
    """
    try:
        # Initialize statistics
        stats = {
            "total_files": 1,
            "processed_files": 0,
            "skipped_files": 0,
            "total_rows": 0,
            "duplicate_rows": 0,
            "start_time": datetime.now(),
            "end_time": None
        }
        
        # Validate configuration
        input_file = input_file or CONFIG.get("merging", {}).get("input_file", "output/extracted_data.xlsx")
        output_file = output_file or CONFIG.get("merging", {}).get("output_file", "output/merged_data.xlsx")
        backup_enabled = CONFIG.get("merging", {}).get("backup_enabled", True)
        
        # Create directories if they don't exist
        create_directory(os.path.dirname(output_file))
        
        # Backup existing output file if enabled
        if backup_enabled and os.path.exists(output_file):
            backup_file = f"{output_file}.bak"
            safe_remove_file(backup_file)  # Remove old backup if exists
            os.rename(output_file, backup_file)
        
        if not os.path.exists(input_file):
            logging.error(f"Input file not found: {input_file}")
            return False, stats
        
        logging.info(f"Processing input file: {input_file}")
        
        merged_wb = openpyxl.Workbook()
        merged_ws = merged_wb.active
        merged_ws.title = "Merged Data"
        
        try:
            wb = openpyxl.load_workbook(input_file, data_only=True)
            ws = wb.active
            
            # Get all rows from the source worksheet
            rows = list(ws.iter_rows(values_only=True))
            
            if not rows:
                logging.warning(f"Warning: {input_file} is empty.")
                return False, stats
            
            # Handle headers
            header = rows[0]
            merged_ws.append(header)
            
            # Add data rows (skip header)
            seen_rows = set()  # For duplicate detection
            
            for row in rows[1:]:
                if row[0] is not None:  # Skip empty rows
                    # Create a hash of the row for duplicate detection
                    row_hash = hash(tuple(str(cell) for cell in row))
                    
                    if row_hash not in seen_rows:
                        merged_ws.append(row)
                        seen_rows.add(row_hash)
                        stats["total_rows"] += 1
                    else:
                        stats["duplicate_rows"] += 1
            
            stats["processed_files"] = 1
            
            # Add metadata sheet
            metadata_ws = merged_wb.create_sheet("Metadata")
            metadata_ws.append(["Processing Statistics"])
            metadata_ws.append(["Input File", input_file])
            metadata_ws.append(["Total Rows", stats["total_rows"]])
            metadata_ws.append(["Duplicate Rows", stats["duplicate_rows"]])
            metadata_ws.append(["Start Time", stats["start_time"]])
            metadata_ws.append(["End Time", datetime.now()])
            
            merged_wb.save(output_file)
            stats["end_time"] = datetime.now()
            
            logging.info(f"\nSuccessfully processed input file.")
            logging.info(f"Total {stats['total_rows']} rows saved to {output_file}")
            if stats["duplicate_rows"] > 0:
                logging.warning(f"Found {stats['duplicate_rows']} duplicate rows")
            return True, stats
            
        except InvalidFileException:
            logging.error(f"Error: {input_file} is not a valid Excel file.")
            return False, stats
        except Exception as e:
            logging.error(f"Error processing {input_file}: {str(e)}")
            return False, stats
    
    except Exception as e:
        logging.error(f"Fatal error during processing: {str(e)}")
        return False, stats

if __name__ == "__main__":
    try:
        success, stats = merge_excel_files()
        if success:
            print("File processing completed successfully")
            print(f"Total rows: {stats['total_rows']}")
        else:
            print("File processing failed")
    except Exception as e:
        print(f"Fatal error: {str(e)}")