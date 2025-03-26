import openpyxl
import re
from config import CONFIG
from utils.helpers import create_directory, safe_remove_file
import os
from typing import List, Set, Optional, Dict
import logging
from tqdm import tqdm

def clean_data(input_file: Optional[str] = None, output_file: Optional[str] = None) -> bool:
    """
    Enhanced data cleaning functionality with:
    - Better stopword handling
    - Improved keyword extraction
    - Robust error handling
    - Config validation
    - Progress reporting
    
    Args:
        input_file (Optional[str]): Path to input Excel file
        output_file (Optional[str]): Path to output Excel file
    
    Returns:
        bool: True if cleaning was successful, False otherwise
    """
    try:
        # Validate and get configuration with defaults
        config_cleaning = CONFIG.get("cleaning", {})
        
        input_file = input_file or config_cleaning.get("input_file", "output/extracted_data.xlsx")
        output_file = output_file or config_cleaning.get("output_file", "output/cleaned_data.xlsx")
        stopwords = set(config_cleaning.get("stopwords", []))
        min_keyword_length = config_cleaning.get("min_keyword_length", 4)
        backup_enabled = config_cleaning.get("backup_enabled", True)
        
        # Create output directory if needed
        create_directory(os.path.dirname(output_file))
        
        # Backup existing output file if enabled
        if backup_enabled and os.path.exists(output_file):
            backup_file = f"{output_file}.bak"
            safe_remove_file(backup_file)  # Remove old backup if exists
            os.rename(output_file, backup_file)
        
        logging.info("Loading workbook...")
        try:
            wb = openpyxl.load_workbook(input_file)
            ws = wb.active
        except FileNotFoundError:
            raise Exception(f"Input file not found: {input_file}")
        except Exception as e:
            raise Exception(f"Error loading workbook: {str(e)}")

        # Validate worksheet structure
        if ws.max_row < 2:
            logging.warning("No data found to clean (empty worksheet).")
            return False
            
        # Add new columns for cleaned data if they don't exist
        headers = {
            4: "Cleaned Content",
            5: "Keywords",
            6: "Word Count",
            7: "Unique Words"
        }
        
        for col, header in headers.items():
            if ws.cell(row=1, column=col).value != header:
                ws.cell(row=1, column=col, value=header)
        
        logging.info(f"Processing {ws.max_row - 1} rows...")
        processed_rows = 0
        skipped_rows = 0
        
        # Create progress bar
        pbar = tqdm(range(2, ws.max_row + 1), desc="Cleaning data")
        
        for row in pbar:
            content = ws.cell(row=row, column=2).value
            if not content:
                skipped_rows += 1
                continue
                
            try:
                # Clean and normalize content
                cleaned_content = clean_text(content, stopwords)
                
                # Extract keywords
                keywords = extract_keywords(cleaned_content, stopwords, min_keyword_length)
                
                # Calculate statistics
                word_count = len(cleaned_content.split())
                unique_words = len(set(cleaned_content.split()))
                
                # Update cells
                ws.cell(row=row, column=4, value=cleaned_content)
                ws.cell(row=row, column=5, value=", ".join(keywords))
                ws.cell(row=row, column=6, value=word_count)
                ws.cell(row=row, column=7, value=unique_words)
                
                processed_rows += 1
                pbar.set_description(f"Processed {processed_rows} rows")
                    
            except Exception as e:
                logging.error(f"Error processing row {row}: {str(e)}")
                skipped_rows += 1
                continue
        
        # Sort data
        logging.info("Sorting data...")
        sort_worksheet(ws)
        
        # Save results
        logging.info("Saving results...")
        wb.save(output_file)
        
        # Log summary
        logging.info(f"Successfully cleaned and saved {processed_rows} rows to {output_file}")
        if skipped_rows > 0:
            logging.warning(f"Skipped {skipped_rows} rows due to errors or empty content")
        
        return True
        
    except Exception as e:
        logging.error(f"Error in clean_data: {str(e)}")
        return False

def clean_text(text: str, stopwords: Set[str]) -> str:
    """
    Clean and normalize text content
    
    Args:
        text (str): Text to clean
        stopwords (Set[str]): Set of stopwords to remove
    
    Returns:
        str: Cleaned text
    """
    if not text:
        return ""
        
    # Convert to string and normalize whitespace
    text = str(text).strip()
    text = re.sub(r'\s+', ' ', text)
    
    # Remove special characters but keep spaces
    text = re.sub(r'[^\w\s]', ' ', text)
    
    # Remove stopwords and short words
    words = [
        word.lower() for word in re.findall(r'\b\w+\b', text)
        if word.lower() not in stopwords and len(word) > 2
    ]
    
    return ' '.join(words)

def extract_keywords(text: str, stopwords: Set[str], min_length: int = 4) -> List[str]:
    """
    Extract meaningful keywords from text
    
    Args:
        text (str): Text to extract keywords from
        stopwords (Set[str]): Set of stopwords to exclude
        min_length (int): Minimum word length for keywords
    
    Returns:
        List[str]: List of keywords
    """
    if not text:
        return []
        
    words = re.findall(r'\b\w+\b', text.lower())
    
    # Get unique words that meet criteria
    keywords = {
        word for word in words
        if (len(word) >= min_length and 
            word not in stopwords and
            not word.isdigit())
    }
    
    return sorted(keywords)

def sort_worksheet(ws) -> None:
    """
    Sort worksheet by first column (Issue number)
    
    Args:
        ws: Worksheet to sort
    """
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            # Try to convert to int for proper numeric sorting
            key = int(row[0]) if str(row[0]).isdigit() else float('inf')
        except (ValueError, TypeError):
            key = float('inf')
        data.append((key, row))
    
    # Sort by the numeric key
    data.sort(key=lambda x: x[0])
    
    # Clear and rewrite sorted data
    ws.delete_rows(2, ws.max_row)
    for _, row in data:
        ws.append(row)

if __name__ == "__main__":
    try:
        success = clean_data()
        if success:
            print("Data cleaning completed successfully")
        else:
            print("Data cleaning failed")
    except Exception as e:
        print(f"Fatal error: {str(e)}")